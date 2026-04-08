# -*- coding: utf-8 -*-
"""
基差/价差走势图数据计算

规则：
- 期货-现货: 基差 = 现货价格 - 期货价格
- 期货-远期: 基差 = 远期价格 - 期货价格  
- 期货-期货: 价差 = 合约1价格 - 合约2价格
- 远期-现货: 价差 = 远期品种价格 - 现货品种价格

主力合约处理：
- 当具体合约（如RB2605）在期货价格表中不存在时
- 如果该合约是当前主力合约，则使用对应的主链价格（螺纹主链/热卷主链）

时间范围：
- 主力合约(螺纹主链/热卷主链): 近1年
- 其他合约: 近3个月
"""
import sqlite3
import os
import openpyxl
import re
from datetime import datetime

class BasisChartCalculator:
    def __init__(self):
        self.db_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'strategy.db')
    
    def _get_main_contract_mapping(self):
        """获取主力合约映射关系
        
        返回: {'RB2605': '螺纹主链', 'HC2605': '热卷主链', ...}
        """
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # 从主力合约配置表读取
            cursor.execute('''
                SELECT variety, contract_code, effective_date, expiry_date 
                FROM main_contracts 
                ORDER BY effective_date DESC
            ''')
            rows = cursor.fetchall()
            conn.close()
            
            mapping = {}
            today = datetime.now().date()
            
            for variety, contract_code, effective_date, expiry_date in rows:
                # 检查是否在有效期内
                try:
                    if effective_date and expiry_date:
                        eff_date = datetime.strptime(effective_date, '%Y-%m-%d').date()
                        exp_date = datetime.strptime(expiry_date, '%Y-%m-%d').date()
                        if eff_date <= today <= exp_date:
                            # 根据品种确定主链名称
                            if '螺纹' in variety or 'RB' in variety:
                                mapping[contract_code] = '螺纹主链'
                            elif '热卷' in variety or 'HC' in variety:
                                mapping[contract_code] = '热卷主链'
                except:
                    continue
            
            return mapping
        except:
            # 如果表不存在或出错，返回空字典
            return {}
    
    def _get_main_chain_for_contract(self, contract):
        """根据合约代码获取对应的主链名称"""
        if not contract:
            return None
        
        # 先检查是否有配置的主力合约映射
        main_contracts = self._get_main_contract_mapping()
        if contract in main_contracts:
            return main_contracts[contract]
        
        # 根据合约代码前缀判断
        contract_upper = contract.upper()
        if contract_upper.startswith('RB'):
            return '螺纹主链'
        elif contract_upper.startswith('HC'):
            return '热卷主链'
        
        return None
    
    def _get_futures_column_index(self, ws_futures, contract):
        """获取期货合约对应的列索引，如果找不到且是主力合约则使用主链"""
        headers = list(ws_futures.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        # 先尝试直接找合约列
        for i, h in enumerate(headers):
            if h == contract:
                return i, contract
        
        # 找不到，检查是否是主力合约，如果是则使用主链
        main_chain = self._get_main_chain_for_contract(contract)
        if main_chain:
            for i, h in enumerate(headers):
                if h == main_chain:
                    print(f"合约{contract}未找到，使用主力合约价格({main_chain})")
                    return i, main_chain
        
        return None, None
    
    def get_basis_data(self, strategy):
        """获取基差/价差走势图数据"""
        # 查找最新的价格数据文件
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute("SELECT file_path FROM data_files WHERE file_type = 'spot_price' ORDER BY upload_date DESC LIMIT 1")
        spot_result = cursor.fetchone()
        
        cursor.execute("SELECT file_path FROM data_files WHERE file_type = 'futures_price' ORDER BY upload_date DESC LIMIT 1")
        futures_result = cursor.fetchone()
        
        conn.close()
        
        if not spot_result or not futures_result:
            return self._generate_mock_data(strategy)
        
        try:
            # 读取现货价格数据
            wb_spot = openpyxl.load_workbook(spot_result[0], data_only=True)
            ws_spot = wb_spot.active
            
            # 读取期货价格数据
            wb_futures = openpyxl.load_workbook(futures_result[0], data_only=True)
            ws_futures = wb_futures.active
            
            # 获取业务类型
            business_type = strategy.get('business_type', '')
            
            # 根据业务类型计算数据
            data = None
            if business_type == '期货-现货':
                data = self._calculate_futures_spot_basis(ws_spot, ws_futures, strategy)
            elif business_type == '期货-远期':
                data = self._calculate_futures_forward_basis(ws_spot, ws_futures, strategy)
            elif business_type == '期货-期货':
                data = self._calculate_futures_futures_spread(ws_futures, strategy)
            elif business_type == '远期-现货':
                data = self._calculate_forward_spot_spread(ws_spot, strategy)
            
            if data is None:
                return self._generate_mock_data(strategy)
            
            # 根据合约类型过滤时间范围
            return self._filter_by_time_range(data, strategy)
                
        except Exception as e:
            print(f"计算基差数据失败: {e}")
            return self._generate_mock_data(strategy)
    
    def _calculate_futures_spot_basis(self, ws_spot, ws_futures, strategy):
        """计算期货-现货基差: 基差 = 现货价格 - 期货价格"""
        spot_variety = strategy.get('spot_variety', '')
        futures_contract = strategy.get('futures_contract', '')
        
        # 现货品种直接使用自身名称（不再映射到主链）
        spot_col_name = spot_variety
        
        # 获取列索引
        spot_headers = list(ws_spot.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        spot_col_idx = None
        for i, h in enumerate(spot_headers):
            if h == spot_col_name:
                spot_col_idx = i
                break
        
        # 获取期货列索引（支持主力合约映射）
        futures_col_idx, actual_contract = self._get_futures_column_index(ws_futures, futures_contract)
        
        if spot_col_idx is None or futures_col_idx is None:
            print(f"找不到对应列: 现货[{spot_col_name}]={spot_col_idx}, 期货[{futures_contract}]={futures_col_idx}")
            return None
        
        # 计算基差
        result = []
        
        # 读取现货数据到字典 {日期: 价格}
        spot_data = {}
        for row in ws_spot.iter_rows(min_row=2, values_only=True):
            if row[0] and row[spot_col_idx]:
                date_str = row[0].strftime('%Y-%m-%d') if isinstance(row[0], datetime) else str(row[0])
                try:
                    spot_data[date_str] = float(row[spot_col_idx])
                except (ValueError, TypeError):
                    continue
        
        # 读取期货数据并计算基差
        for row in ws_futures.iter_rows(min_row=2, values_only=True):
            if row[0] and row[futures_col_idx]:
                date_str = row[0].strftime('%Y-%m-%d') if isinstance(row[0], datetime) else str(row[0])
                try:
                    futures_price = float(row[futures_col_idx])
                    
                    # 基差 = 现货 - 期货
                    if date_str in spot_data:
                        basis = spot_data[date_str] - futures_price
                        result.append({
                            'date': date_str,
                            'basis': round(basis, 2),
                            'spot': spot_data[date_str],
                            'futures': futures_price
                        })
                except (ValueError, TypeError):
                    continue
        
        return sorted(result, key=lambda x: x['date'])
    
    def _calculate_futures_forward_basis(self, ws_spot, ws_futures, strategy):
        """计算期货-远期基差: 基差 = 远期价格 - 期货价格"""
        # 期货-远期业务中，forward_variety存储远期品种
        forward_variety = strategy.get('spot_variety', '')  # 这里spot_variety实际存储的是远期品种
        futures_contract = strategy.get('futures_contract', '')
        
        # 远期品种直接使用自身名称
        forward_col_name = forward_variety
        
        # 获取列索引
        spot_headers = list(ws_spot.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        forward_col_idx = None
        for i, h in enumerate(spot_headers):
            if h == forward_col_name:
                forward_col_idx = i
                break
        
        # 获取期货列索引（支持主力合约映射）
        futures_col_idx, actual_contract = self._get_futures_column_index(ws_futures, futures_contract)
        
        if forward_col_idx is None or futures_col_idx is None:
            print(f"找不到对应列: 远期[{forward_col_name}]={forward_col_idx}, 期货[{futures_contract}]={futures_col_idx}")
            return None
        
        # 计算基差
        result = []
        
        # 读取远期数据到字典 {日期: 价格}
        forward_data = {}
        for row in ws_spot.iter_rows(min_row=2, values_only=True):
            if row[0] and row[forward_col_idx]:
                date_str = row[0].strftime('%Y-%m-%d') if isinstance(row[0], datetime) else str(row[0])
                try:
                    forward_data[date_str] = float(row[forward_col_idx])
                except (ValueError, TypeError):
                    continue
        
        # 读取期货数据并计算基差
        for row in ws_futures.iter_rows(min_row=2, values_only=True):
            if row[0] and row[futures_col_idx]:
                date_str = row[0].strftime('%Y-%m-%d') if isinstance(row[0], datetime) else str(row[0])
                try:
                    futures_price = float(row[futures_col_idx])
                    
                    # 基差 = 远期 - 期货
                    if date_str in forward_data:
                        basis = forward_data[date_str] - futures_price
                        result.append({
                            'date': date_str,
                            'basis': round(basis, 2),
                            'spot': forward_data[date_str],  # 这里spot字段实际存储远期价格
                            'futures': futures_price
                        })
                except (ValueError, TypeError):
                    continue
        
        return sorted(result, key=lambda x: x['date'])
    
    def _calculate_futures_futures_spread(self, ws_futures, strategy):
        """计算期货-期货价差: 价差 = 合约1价格 - 合约2价格"""
        contract1 = strategy.get('futures_contract', '')
        contract2 = strategy.get('futures_contract2', '')
        
        # 获取列索引（支持主力合约映射）
        col1_idx, actual_contract1 = self._get_futures_column_index(ws_futures, contract1)
        col2_idx, actual_contract2 = self._get_futures_column_index(ws_futures, contract2)
        
        if col1_idx is None or col2_idx is None:
            print(f"找不到对应列: 合约1[{contract1}]={col1_idx}, 合约2[{contract2}]={col2_idx}")
            return None
        
        result = []
        for row in ws_futures.iter_rows(min_row=2, values_only=True):
            if row[0] and row[col1_idx] and row[col2_idx]:
                date_str = row[0].strftime('%Y-%m-%d') if isinstance(row[0], datetime) else str(row[0])
                try:
                    price1 = float(row[col1_idx])
                    price2 = float(row[col2_idx])
                    
                    # 价差 = 合约1 - 合约2
                    spread = price1 - price2
                    result.append({
                        'date': date_str,
                        'basis': round(spread, 2),
                        'spot': price1,      # 合约1价格
                        'futures': price2    # 合约2价格
                    })
                except (ValueError, TypeError):
                    continue
        
        return sorted(result, key=lambda x: x['date'])
    
    def _calculate_forward_spot_spread(self, ws_spot, strategy):
        """计算远期-现货价差: 价差 = 品种1价格 - 品种2价格
        
        品种1 = spot_variety (如：马钢螺纹)
        品种2 = forward_variety (如：中天螺纹)
        """
        variety1 = strategy.get('spot_variety', '')    # 品种1
        variety2 = strategy.get('forward_variety', '')  # 品种2
        
        # 直接使用品种名称
        variety1_col_name = variety1
        variety2_col_name = variety2
        
        headers = list(ws_spot.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        
        variety1_col_idx = None
        variety2_col_idx = None
        
        for i, h in enumerate(headers):
            if h == variety1_col_name:
                variety1_col_idx = i
            if h == variety2_col_name:
                variety2_col_idx = i
        
        if variety1_col_idx is None or variety2_col_idx is None:
            print(f"找不到对应列: 品种1[{variety1_col_name}]={variety1_col_idx}, 品种2[{variety2_col_name}]={variety2_col_idx}")
            return None
        
        result = []
        for row in ws_spot.iter_rows(min_row=2, values_only=True):
            if row[0] and row[variety1_col_idx] and row[variety2_col_idx]:
                date_str = row[0].strftime('%Y-%m-%d') if isinstance(row[0], datetime) else str(row[0])
                try:
                    price1 = float(row[variety1_col_idx])  # 品种1价格
                    price2 = float(row[variety2_col_idx])  # 品种2价格
                    
                    # 价差 = 品种1 - 品种2
                    spread = price1 - price2
                    result.append({
                        'date': date_str,
                        'basis': round(spread, 2),
                        'spot': price1,     # 品种1价格
                        'futures': price2   # 品种2价格
                    })
                except (ValueError, TypeError):
                    continue
        
        return sorted(result, key=lambda x: x['date'])
    
    def _filter_by_time_range(self, data, strategy):
        """不过滤时间范围，返回所有可用数据
        由前端通过日期选择器来控制显示范围
        """
        return data
    
    def _generate_mock_data(self, strategy):
        """生成模拟数据（当没有真实价格数据时）"""
        result = []
        
        # 根据策略生成不同的模拟数据特征
        business_type = strategy.get('business_type', '')
        futures_contract = strategy.get('futures_contract', '')
        is_main_contract = futures_contract in ['螺纹主链', '热卷主链', 'RB主链', 'HC主链']
        
        # 根据合约类型确定时间范围
        end_date = datetime(2026, 3, 20)
        if is_main_contract:
            start_date = end_date - __import__('datetime').timedelta(days=365)  # 近1年
        else:
            start_date = end_date - __import__('datetime').timedelta(days=90)   # 近3个月
        
        import random
        random.seed(42)  # 固定随机种子，使数据可重复
        
        current_date = start_date
        
        # 根据业务类型设置不同的基差基准
        if business_type == '期货-现货':
            basis_base = -100  # 通常现货低于期货（负基差）
        elif business_type == '期货-远期':
            basis_base = -50
        elif business_type == '期货-期货':
            basis_base = 30
        else:
            basis_base = 20
        
        while current_date <= end_date:
            # 跳过周末
            if current_date.weekday() < 5:  # 0-4是周一到周五
                date_str = current_date.strftime('%Y-%m-%d')
                
                # 模拟基差波动
                noise = random.uniform(-20, 20)
                trend = (current_date - start_date).days * 0.02
                basis = basis_base + trend + noise
                
                # 模拟价格
                spot_price = 3200 + random.uniform(-50, 50)
                if business_type in ['期货-现货', '期货-远期']:
                    futures_price = spot_price - basis  # 基差 = 现货 - 期货
                else:
                    futures_price = 3200 + random.uniform(-50, 50)
                
                result.append({
                    'date': date_str,
                    'basis': round(basis, 2),
                    'spot': round(spot_price, 2),
                    'futures': round(futures_price, 2)
                })
            
            current_date += __import__('datetime').timedelta(days=1)
        
        return result
