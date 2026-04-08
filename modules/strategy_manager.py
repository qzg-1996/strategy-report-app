# -*- coding: utf-8 -*-
"""
策略管理器 - 管理策略的增删改查和盈亏计算

持仓盈亏计算规则：

【期货-现货 / 期货-远期】
- 基差 = 现货价格 - 期货价格
- 期货方向=空: 持仓盈亏 = (期末基差 - 建仓基差) × 建仓手数 × 10
- 期货方向=多: 持仓盈亏 = (建仓基差 - 期末基差) × 建仓手数 × 10

【期货-期货】
- 逐行计算，只有期货端
- 期货方向=空: 持仓盈亏 = (期货建仓价 - 期货最新价) × 建仓手数 × 10
- 期货方向=多: 持仓盈亏 = (期货最新价 - 期货建仓价) × 建仓手数 × 10
- 全部求和

【远期-现货】
- 看品种1的方向（策略中的 spot_variety）
- 品种1方向=卖:
  持仓盈亏 = (品种1建仓价 - 品种1最新价) × 品种1吨 + (品种2最新价 - 品种2建仓价) × 品种2吨
- 品种1方向=买:
  持仓盈亏 = (品种1最新价 - 品种1建仓价) × 品种1吨 + (品种2建仓价 - 品种2最新价) × 品种2吨

【投机】
- 期货方向=空: 持仓盈亏 = (期货建仓价 - 期货最新价) × 建仓手数 × 10
- 期货方向=多: 持仓盈亏 = (期货最新价 - 期货建仓价) × 建仓手数 × 10

主力合约处理：
- 当具体合约（如RB2605）在期货价格表中不存在时
- 如果该合约是当前主力合约，则使用对应的主链价格（螺纹主链/热卷主链）
"""
import sqlite3
import os
import openpyxl
from datetime import datetime

class StrategyManager:
    def __init__(self):
        self.db_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'strategy.db')
    
    def _get_main_contract_mapping(self):
        """获取主力合约映射关系
        
        返回: {'RB2605': '螺纹主链', 'HC2605': '热卷主链', ...}
        """
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # 从主力合约配置表读取
            cursor.execute('''
                SELECT variety, contract_code, effective_date, expiry_date 
                FROM main_contracts 
                ORDER BY effective_date DESC
            ''')
            rows = cursor.fetchall()
            conn.close()
            
            mapping = {}
            today = datetime.now().date()
            
            for variety, contract_code, effective_date, expiry_date in rows:
                # 检查是否在有效期内
                try:
                    if effective_date and expiry_date:
                        eff_date = datetime.strptime(effective_date, '%Y-%m-%d').date()
                        exp_date = datetime.strptime(expiry_date, '%Y-%m-%d').date()
                        if eff_date <= today <= exp_date:
                            # 根据品种确定主链名称
                            if '螺纹' in variety or 'RB' in variety:
                                mapping[contract_code] = '螺纹主链'
                            elif '热卷' in variety or 'HC' in variety:
                                mapping[contract_code] = '热卷主链'
                except:
                    continue
            
            return mapping
        except:
            # 如果表不存在或出错，返回空字典
            return {}
    
    def _get_main_chain_for_contract(self, contract):
        """根据合约代码获取对应的主链名称"""
        if not contract:
            return None
        
        # 先检查是否有配置的主力合约映射
        main_contracts = self._get_main_contract_mapping()
        if contract in main_contracts:
            return main_contracts[contract]
        
        # 根据合约代码前缀判断
        contract_upper = contract.upper()
        if contract_upper.startswith('RB'):
            return '螺纹主链'
        elif contract_upper.startswith('HC'):
            return '热卷主链'
        
        return None
    
    def get_summary_data(self):
        """获取策略汇总数据 - 优化版：只读取一次Excel"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM strategies ORDER BY sort_order, id')
        columns = [description[0] for description in cursor.description]
        strategies = [dict(zip(columns, row)) for row in cursor.fetchall()]
        
        # 读取交易记录文件
        cursor.execute("SELECT file_path FROM data_files WHERE file_type = 'trade_record' ORDER BY upload_date DESC LIMIT 1")
        trade_result = cursor.fetchone()
        
        # 读取远期-现货交易记录文件
        cursor.execute("SELECT file_path FROM data_files WHERE file_type = 'forward_spot_record' ORDER BY upload_date DESC LIMIT 1")
        forward_spot_result = cursor.fetchone()
        
        conn.close()
        
        # 缓存交易记录数据
        all_trade_records = []
        if trade_result and os.path.exists(trade_result[0]):
            try:
                wb = openpyxl.load_workbook(trade_result[0], data_only=True)
                ws = wb.active
                all_trade_records = list(ws.iter_rows(min_row=3, values_only=True))
            except Exception as e:
                print(f"读取交易记录失败: {e}")
        
        # 缓存远期-现货交易记录数据
        all_forward_spot_records = []
        if forward_spot_result and os.path.exists(forward_spot_result[0]):
            try:
                wb = openpyxl.load_workbook(forward_spot_result[0], data_only=True)
                ws = wb.active
                all_forward_spot_records = list(ws.iter_rows(min_row=4, values_only=True))  # 第4行开始
            except Exception as e:
                print(f"读取远期-现货交易记录失败: {e}")
        
        result = []
        total_position = 0
        total_position_profit = 0
        total_close = 0
        total_close_profit = 0
        
        for strategy in strategies:
            business_type = strategy.get('business_type', '')
            
            # 远期-现货使用专门的记录
            if business_type == '远期-现货':
                summary = self._calculate_forward_spot_summary_optimized(strategy, all_forward_spot_records)
            else:
                summary = self.calculate_strategy_summary_optimized(strategy, all_trade_records)
            
            result.append({
                'id': strategy['id'],
                'strategy_code': strategy['strategy_code'],
                'strategy_name': strategy['strategy_name'],
                'strategy_type': strategy['strategy_type'],
                'business_type': strategy['business_type'],
                'plan_quantity': strategy['plan_quantity'] or 0,
                'position_quantity': summary['position_quantity'],
                'position_profit': summary['position_profit'],
                'close_quantity': summary['close_quantity'],
                'close_profit': summary['close_profit']
            })
            total_position += summary['position_quantity']
            total_position_profit += summary['position_profit']
            total_close += summary['close_quantity']
            total_close_profit += summary['close_profit']
        
        return {
            'strategies': result,
            'total_position': total_position,
            'total_position_profit': total_position_profit,
            'total_close': total_close,
            'total_close_profit': total_close_profit
        }
    
    def calculate_strategy_summary_optimized(self, strategy, all_trade_records):
        """计算单个策略的汇总数据 - 优化版：使用缓存的交易记录（期货类策略）"""
        total_futures_lots = 0
        total_close_lots = 0
        total_close_profit = 0
        position_records = []
        
        strategy_type = strategy.get('strategy_type', '')
        business_type = strategy.get('business_type', '')
        futures_direction = strategy.get('futures_direction', '')
        direction_factor = -1 if futures_direction == '空' else 1
        
        # 远期-现货策略应该使用专门的方法
        if business_type == '远期-现货':
            return {'position_quantity': 0, 'position_profit': 0, 'close_quantity': 0, 'close_profit': 0}
        
        for row in all_trade_records:
            if not row[1] or str(row[1]) != strategy_type:
                continue
            
            # 匹配交易记录
            if not self._match_trade_record(row, strategy, business_type):
                continue
            
            futures_lots = float(row[7]) if row[7] else 0
            close_lots = float(row[14]) if row[14] else 0
            futures_open_price = float(row[8]) if row[8] else 0
            futures_close_price = float(row[15]) if row[15] else 0
            spot_lots = float(row[23]) if row[23] else 0
            spot_open_price = float(row[24]) if row[24] else 0
            
            total_futures_lots += futures_lots
            total_close_lots += close_lots
            
            # 计算平仓盈亏（期货端）
            if close_lots > 0 and futures_close_price > 0:
                close_pnl = (futures_close_price - futures_open_price) * close_lots * 10 * direction_factor
                total_close_profit += close_pnl
            
            # 记录持仓明细
            remaining = futures_lots - close_lots
            if remaining > 0:
                record = {
                    'lots': remaining,
                    'futures_open_price': futures_open_price,
                    'spot_open_price': spot_open_price,
                    'direction_factor': direction_factor,
                    'futures_direction': futures_direction
                }
                position_records.append(record)
        
        # 手数转吨位
        position_lots = total_futures_lots - total_close_lots
        position_tons = position_lots * 10
        close_tons = total_close_lots * 10
        
        # 计算持仓盈亏
        position_profit = self._calculate_position_profit(strategy, position_records)
        
        return {
            'position_quantity': round(position_tons, 2),
            'position_profit': round(position_profit, 2),
            'close_quantity': round(close_tons, 2),
            'close_profit': round(total_close_profit, 2)
        }
    
    def _calculate_forward_spot_summary_optimized(self, strategy, all_forward_spot_records):
        """计算远期-现货策略汇总 - 优化版：使用缓存的交易记录"""
        total_open_tons = 0
        total_close_tons = 0
        total_close_profit = 0
        position_records = []
        
        strategy_type = strategy.get('strategy_type', '')
        futures_direction = strategy.get('futures_direction', '')
        
        for row in all_forward_spot_records:
            if not row[0] or str(row[0]) != strategy_type:
                continue
            
            # 匹配交易记录
            if not self._match_trade_record(row, strategy, '远期-现货'):
                continue
            
            # 读取数据
            variety1_open_tons = float(row[8]) if row[8] else 0
            variety1_open_price = float(row[7]) if row[7] else 0
            variety2_open_price = float(row[13]) if row[13] else 0
            
            variety1_close_tons = float(row[23]) if len(row) > 23 and row[23] else 0
            variety1_close_price = float(row[22]) if len(row) > 22 and row[22] else 0
            variety2_close_price = float(row[28]) if len(row) > 28 and row[28] else 0
            
            total_open_tons += variety1_open_tons
            total_close_tons += variety1_close_tons
            
            # 计算平仓盈亏
            if variety1_close_tons > 0:
                open_spread = variety1_open_price - variety2_open_price
                close_spread = variety1_close_price - variety2_close_price
                spread_change = close_spread - open_spread
                
                if futures_direction == '空':
                    close_pnl = -spread_change * variety1_close_tons
                else:
                    close_pnl = spread_change * variety1_close_tons
                
                total_close_profit += close_pnl
            
            # 记录持仓明细
            remaining_tons = variety1_open_tons - variety1_close_tons
            if remaining_tons > 0:
                position_records.append({
                    'variety1_tons': remaining_tons,
                    'variety1_open_price': variety1_open_price,
                    'variety2_open_price': variety2_open_price,
                    'direction': futures_direction
                })
        
        # 计算持仓盈亏
        position_profit = self._calculate_forward_spot_position_profit(strategy, position_records)
        
        return {
            'position_quantity': round(total_open_tons - total_close_tons, 2),
            'position_profit': round(position_profit, 2),
            'close_quantity': round(total_close_tons, 2),
            'close_profit': round(total_close_profit, 2)
        }
    
    def calculate_strategy_summary(self, strategy):
        """计算单个策略的汇总数据"""
        business_type = strategy.get('business_type', '')
        
        # 远期-现货使用专门的计算方法
        if business_type == '远期-现货':
            return self._calculate_forward_spot_summary(strategy)
        
        # 其他类型使用原有方法
        return self._calculate_futures_based_summary(strategy)
    
    def _calculate_futures_based_summary(self, strategy):
        """计算基于期货的策略汇总（期货-现货、期货-远期、期货-期货、投机）"""
        # 获取交易记录文件
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT file_path FROM data_files WHERE file_type = 'trade_record' ORDER BY upload_date DESC LIMIT 1")
        result = cursor.fetchone()
        conn.close()
        
        total_futures_lots = 0
        total_close_lots = 0
        total_close_profit = 0
        position_records = []
        
        if result and os.path.exists(result[0]):
            try:
                wb = openpyxl.load_workbook(result[0], data_only=True)
                ws = wb.active
                
                strategy_type = strategy.get('strategy_type', '')
                business_type = strategy.get('business_type', '')
                futures_direction = strategy.get('futures_direction', '')
                direction_factor = -1 if futures_direction == '空' else 1
                
                for row in ws.iter_rows(min_row=3, values_only=True):
                    if not row[1] or str(row[1]) != strategy_type:
                        continue
                    
                    # 匹配交易记录
                    if not self._match_trade_record(row, strategy, business_type):
                        continue
                    
                    futures_lots = float(row[7]) if row[7] else 0
                    close_lots = float(row[14]) if row[14] else 0
                    futures_open_price = float(row[8]) if row[8] else 0
                    futures_close_price = float(row[15]) if row[15] else 0
                    spot_lots = float(row[23]) if row[23] else 0
                    spot_open_price = float(row[24]) if row[24] else 0
                    
                    total_futures_lots += futures_lots
                    total_close_lots += close_lots
                    
                    # 计算平仓盈亏（期货端）
                    if close_lots > 0 and futures_close_price > 0:
                        close_pnl = (futures_close_price - futures_open_price) * close_lots * 10 * direction_factor
                        total_close_profit += close_pnl
                    
                    # 记录持仓明细
                    remaining = futures_lots - close_lots
                    if remaining > 0:
                        record = {
                            'lots': remaining,
                            'futures_open_price': futures_open_price,
                            'spot_open_price': spot_open_price,
                            'direction_factor': direction_factor,
                            'futures_direction': futures_direction
                        }
                        position_records.append(record)
                        
            except Exception as e:
                print(f"读取交易记录失败: {e}")
        
        # 手数转吨位
        position_lots = total_futures_lots - total_close_lots
        position_tons = position_lots * 10
        close_tons = total_close_lots * 10
        
        # 计算持仓盈亏
        position_profit = self._calculate_position_profit(strategy, position_records)
        
        return {
            'position_quantity': round(position_tons, 2),
            'position_profit': round(position_profit, 2),
            'close_quantity': round(close_tons, 2),
            'close_profit': round(total_close_profit, 2)
        }
    
    def _calculate_forward_spot_summary(self, strategy):
        """计算远期-现货策略汇总（两个现货品种价差交易）"""
        # 获取远期-现货交易记录文件
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT file_path FROM data_files WHERE file_type = 'forward_spot_record' ORDER BY upload_date DESC LIMIT 1")
        result = cursor.fetchone()
        conn.close()
        
        total_open_tons = 0  # 总开仓吨数
        total_close_tons = 0  # 总平仓吨数
        total_close_profit = 0  # 总平仓盈亏
        position_records = []  # 持仓明细
        
        if result and os.path.exists(result[0]):
            try:
                wb = openpyxl.load_workbook(result[0], data_only=True)
                ws = wb.active
                
                strategy_type = strategy.get('strategy_type', '')
                futures_direction = strategy.get('futures_direction', '')  # 品种1的方向
                
                # 遍历数据（从第4行开始，第3行是表头）
                for row in ws.iter_rows(min_row=4, values_only=True):
                    if not row[0] or str(row[0]) != strategy_type:  # 第0列是策略类型
                        continue
                    
                    # 匹配交易记录
                    if not self._match_trade_record(row, strategy, '远期-现货'):
                        continue
                    
                    # 读取数据
                    # 品种1：建仓
                    variety1_open_tons = float(row[8]) if row[8] else 0  # 吨数1
                    variety1_open_price = float(row[7]) if row[7] else 0  # 价格1
                    
                    # 品种2：建仓
                    variety2_open_tons = float(row[14]) if row[14] else 0  # 吨数2
                    variety2_open_price = float(row[13]) if row[13] else 0  # 价格2
                    
                    # 平仓数据（如果有）
                    variety1_close_tons = float(row[23]) if len(row) > 23 and row[23] else 0
                    variety1_close_price = float(row[22]) if len(row) > 22 and row[22] else 0
                    variety2_close_tons = float(row[29]) if len(row) > 29 and row[29] else 0
                    variety2_close_price = float(row[28]) if len(row) > 28 and row[28] else 0
                    
                    total_open_tons += variety1_open_tons
                    total_close_tons += variety1_close_tons
                    
                    # 计算平仓盈亏
                    if variety1_close_tons > 0:
                        # 价差变化
                        open_spread = variety1_open_price - variety2_open_price
                        close_spread = variety1_close_price - variety2_close_price
                        spread_change = close_spread - open_spread
                        
                        # 品种1方向为"空"：卖出品种1、买入品种2
                        # 盈利 = 价差缩小（卖出价高，买入价低）
                        if futures_direction == '空':
                            close_pnl = -spread_change * variety1_close_tons
                        else:  # 品种1方向为"多"：买入品种1、卖出品种2
                            close_pnl = spread_change * variety1_close_tons
                        
                        total_close_profit += close_pnl
                    
                    # 记录持仓明细
                    remaining_tons = variety1_open_tons - variety1_close_tons
                    if remaining_tons > 0:
                        position_records.append({
                            'variety1_tons': remaining_tons,
                            'variety1_open_price': variety1_open_price,
                            'variety2_open_price': variety2_open_price,
                            'direction': futures_direction
                        })
                        
            except Exception as e:
                print(f"读取远期-现货交易记录失败: {e}")
        
        # 计算持仓盈亏
        position_profit = self._calculate_forward_spot_position_profit(strategy, position_records)
        
        return {
            'position_quantity': round(total_open_tons - total_close_tons, 2),
            'position_profit': round(position_profit, 2),
            'close_quantity': round(total_close_tons, 2),
            'close_profit': round(total_close_profit, 2)
        }
    
    def _calculate_forward_spot_position_profit(self, strategy, position_records):
        """计算远期-现货持仓盈亏"""
        if not position_records:
            return 0
        
        variety1 = strategy.get('spot_variety', '')
        variety2 = strategy.get('forward_variety', '')
        direction = strategy.get('futures_direction', '')
        
        # 获取最新价格
        variety1_info = self._get_latest_spot_price_with_date(variety1)
        if not variety1_info:
            return 0
        
        current_price1 = variety1_info['price']
        latest_date = variety1_info['date']
        
        current_price2 = self._get_spot_price_by_date(variety2, latest_date)
        if current_price2 is None:
            return 0
        
        total_pnl = 0
        for record in position_records:
            open_price1 = record['variety1_open_price']
            open_price2 = record['variety2_open_price']
            tons = record['variety1_tons']
            
            open_spread = open_price1 - open_price2
            current_spread = current_price1 - current_price2
            spread_change = current_spread - open_spread
            
            if direction == '空':  # 卖出品种1、买入品种2
                pnl = -spread_change * tons
            else:  # 买入品种1、卖出品种2
                pnl = spread_change * tons
            
            total_pnl += pnl
        
        return total_pnl
    
    def _calculate_position_profit(self, strategy, position_records):
        """计算持仓盈亏"""
        if not position_records:
            return 0
        
        strategy_type = strategy.get('strategy_type', '')
        business_type = strategy.get('business_type', '')
        futures_direction = strategy.get('futures_direction', '')
        
        # 获取最新价格（以期货最新日期为准，取同一日期的现货价格）
        futures_contract = strategy.get('futures_contract', '')
        spot_variety = strategy.get('spot_variety', '')
        
        total_pnl = 0
        calculation_failed = False
        
        if business_type in ['期货-现货', '期货-远期']:
            # 【期货-现货 / 期货-远期】
            # 基差 = 现货价格 - 期货价格
            # 期货方向=空: 持仓盈亏 = (期末基差 - 建仓基差) × 建仓手数 × 10
            # 期货方向=多: 持仓盈亏 = (建仓基差 - 期末基差) × 建仓手数 × 10
            
            # 获取期货最新价格和日期
            futures_info = self._get_latest_futures_price_with_date(futures_contract)
            if not futures_info:
                print(f"[持仓盈亏计算] 无法获取期货价格: {futures_contract}")
                calculation_failed = True
            else:
                current_futures_price = futures_info['price']
                latest_date = futures_info['date']
                
                # 获取现货/远期最新价格（期货-远期时spot_variety存储的是远期品种）
                current_spot_price = self._get_spot_price_by_date(spot_variety, latest_date)
                if current_spot_price is None:
                    print(f"[持仓盈亏计算] 无法获取现货价格: {spot_variety}, 日期: {latest_date}")
                    calculation_failed = True
                else:
                    for record in position_records:
                        lots = record['lots']
                        
                        # 建仓基差 = 现货建仓价 - 期货建仓价
                        open_basis = record['spot_open_price'] - record['futures_open_price']
                        # 期末基差 = 现货最新价 - 期货最新价
                        current_basis = current_spot_price - current_futures_price
                        
                        # 根据期货方向计算盈亏
                        if futures_direction == '空':
                            # 空: (期末基差 - 建仓基差) × 手数 × 10
                            basis_change = current_basis - open_basis
                        else:  # 多
                            # 多: (建仓基差 - 期末基差) × 手数 × 10
                            basis_change = open_basis - current_basis
                        
                        pnl = lots * 10 * basis_change
                        total_pnl += pnl
                        
        elif business_type == '期货-期货':
            # 【期货-期货】
            # 逐行计算，只有期货端
            # 期货方向=空: 持仓盈亏 = (期货建仓价 - 期货最新价) × 建仓手数 × 10
            # 期货方向=多: 持仓盈亏 = (期货最新价 - 期货建仓价) × 建仓手数 × 10
            
            contract2 = strategy.get('futures_contract2', '')
            
            # 获取两个合约的最新价格
            futures_info1 = self._get_latest_futures_price_with_date(futures_contract)
            futures_info2 = self._get_latest_futures_price_with_date(contract2)
            
            if not futures_info1 or not futures_info2:
                print(f"[持仓盈亏计算] 无法获取期货价格: 合约1={futures_contract}, 合约2={contract2}")
                calculation_failed = True
            else:
                current_price1 = futures_info1['price']
                current_price2 = futures_info2['price']
                
                # 注意：期货-期货的持仓记录中，需要区分是哪个合约的记录
                for record in position_records:
                    lots = record['lots']
                    record_direction = record.get('futures_direction', futures_direction)
                    open_price = record['futures_open_price']
                    
                    # 根据方向计算盈亏
                    if record_direction == '空':
                        # 空: (建仓价 - 最新价) × 手数 × 10
                        price_change = open_price - current_price1
                    else:  # 多
                        # 多: (最新价 - 建仓价) × 手数 × 10
                        price_change = current_price1 - open_price
                    
                    pnl = lots * 10 * price_change
                    total_pnl += pnl
                
        elif business_type == '远期-现货':
            # 【远期-现货】价差交易（两个现货品种）
            # 使用专门的计算方法
            total_pnl = self._calculate_forward_spot_position_profit(strategy, position_records)
                
        elif business_type == '投机' or strategy_type == '趋势交易':
            # 【投机】
            # 期货方向=空: 持仓盈亏 = (期货建仓价 - 期货最新价) × 建仓手数 × 10
            # 期货方向=多: 持仓盈亏 = (期货最新价 - 期货建仓价) × 建仓手数 × 10
            
            # 获取期货最新价格
            futures_info = self._get_latest_futures_price_with_date(futures_contract)
            if not futures_info:
                print(f"[持仓盈亏计算] 无法获取期货价格: {futures_contract}")
                calculation_failed = True
            else:
                current_futures_price = futures_info['price']
                
                for record in position_records:
                    lots = record['lots']
                    record_direction = record.get('futures_direction', futures_direction)
                    open_price = record['futures_open_price']
                    
                    # 根据方向计算盈亏
                    if record_direction == '空':
                        # 空: (建仓价 - 最新价) × 手数 × 10
                        price_change = open_price - current_futures_price
                    else:  # 多
                        # 多: (最新价 - 建仓价) × 手数 × 10
                        price_change = current_futures_price - open_price
                    
                    pnl = lots * 10 * price_change
                    total_pnl += pnl
        
        if calculation_failed and total_pnl == 0:
            print(f"[持仓盈亏计算] 策略 {strategy.get('strategy_name', 'Unknown')} 因价格数据缺失无法计算")
        
        return total_pnl
    
    def _get_latest_futures_price_with_date(self, contract):
        """获取期货最新价格和日期
        
        返回: {'price': float, 'date': datetime}
        """
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT file_path FROM data_files WHERE file_type = 'futures_price' ORDER BY upload_date DESC LIMIT 1")
            result = cursor.fetchone()
            conn.close()
            
            if not result or not os.path.exists(result[0]):
                return None
            
            wb = openpyxl.load_workbook(result[0], data_only=True)
            ws = wb.active
            
            headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            col_idx = None
            actual_contract = contract
            
            for i, h in enumerate(headers):
                if h == contract:
                    col_idx = i
                    break
            
            # 如果没找到，尝试使用主力合约映射
            if col_idx is None:
                main_chain = self._get_main_chain_for_contract(contract)
                if main_chain:
                    for i, h in enumerate(headers):
                        if h == main_chain:
                            col_idx = i
                            actual_contract = main_chain
                            print(f"合约{contract}未找到，使用主力合约价格({main_chain})")
                            break
            
            if col_idx is None:
                return None
            
            # 找最新日期的价格
            latest_price = None
            latest_date = None
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[col_idx]:
                    date_val = row[0]
                    if isinstance(date_val, str):
                        try:
                            date_val = datetime.strptime(date_val, '%Y-%m-%d')
                        except:
                            continue
                    elif isinstance(date_val, datetime):
                        pass  # 已经是datetime类型
                    else:
                        continue
                    
                    if latest_date is None or date_val > latest_date:
                        latest_date = date_val
                        latest_price = float(row[col_idx])
            
            if latest_price is None:
                return None
            
            return {'price': latest_price, 'date': latest_date}
            
        except Exception as e:
            print(f"获取期货最新价格失败: {e}")
            return None
    
    def _get_latest_spot_price_with_date(self, variety):
        """获取现货最新价格和日期
        
        返回: {'price': float, 'date': datetime}
        """
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT file_path FROM data_files WHERE file_type = 'spot_price' ORDER BY upload_date DESC LIMIT 1")
            result = cursor.fetchone()
            conn.close()
            
            if not result or not os.path.exists(result[0]):
                return None
            
            wb = openpyxl.load_workbook(result[0], data_only=True)
            ws = wb.active
            
            headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            col_idx = None
            for i, h in enumerate(headers):
                if h == variety:
                    col_idx = i
                    break
            
            if col_idx is None:
                return None
            
            # 找最新日期的价格
            latest_price = None
            latest_date = None
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[col_idx]:
                    date_val = row[0]
                    if isinstance(date_val, str):
                        try:
                            date_val = datetime.strptime(date_val, '%Y-%m-%d')
                        except:
                            continue
                    elif isinstance(date_val, datetime):
                        pass
                    else:
                        continue
                    
                    if latest_date is None or date_val > latest_date:
                        latest_date = date_val
                        latest_price = float(row[col_idx])
            
            if latest_price is None:
                return None
            
            return {'price': latest_price, 'date': latest_date}
            
        except Exception as e:
            print(f"获取现货最新价格失败: {e}")
            return None
    
    def _get_spot_price_by_date(self, spot_variety, target_date):
        """根据日期获取现货价格"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT file_path FROM data_files WHERE file_type = 'spot_price' ORDER BY upload_date DESC LIMIT 1")
            result = cursor.fetchone()
            conn.close()
            
            if not result or not os.path.exists(result[0]):
                return None
            
            wb = openpyxl.load_workbook(result[0], data_only=True)
            ws = wb.active
            
            # 现货品种直接对应现货表中的列名
            col_name = spot_variety
            
            headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            col_idx = None
            for i, h in enumerate(headers):
                if h == col_name:
                    col_idx = i
                    break
            
            if col_idx is None:
                return None
            
            # 找到对应日期的价格
            target_date_only = target_date.date() if isinstance(target_date, datetime) else target_date
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[col_idx]:
                    date_val = row[0]
                    if isinstance(date_val, str):
                        try:
                            date_val = datetime.strptime(date_val, '%Y-%m-%d').date()
                        except:
                            continue
                    elif isinstance(date_val, datetime):
                        date_val = date_val.date()
                    
                    if date_val == target_date_only:
                        return float(row[col_idx])
            
            return None
            
        except Exception as e:
            print(f"获取现货价格失败: {e}")
            return None
    
    def _get_futures_price_by_date(self, contract, target_date):
        """根据日期获取期货价格"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT file_path FROM data_files WHERE file_type = 'futures_price' ORDER BY upload_date DESC LIMIT 1")
            result = cursor.fetchone()
            conn.close()
            
            if not result or not os.path.exists(result[0]):
                return None
            
            wb = openpyxl.load_workbook(result[0], data_only=True)
            ws = wb.active
            
            headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            col_idx = None
            actual_contract = contract
            
            for i, h in enumerate(headers):
                if h == contract:
                    col_idx = i
                    break
            
            # 如果没找到，尝试使用主力合约映射
            if col_idx is None:
                main_chain = self._get_main_chain_for_contract(contract)
                if main_chain:
                    for i, h in enumerate(headers):
                        if h == main_chain:
                            col_idx = i
                            actual_contract = main_chain
                            print(f"合约{contract}未找到，使用主力合约价格({main_chain})")
                            break
            
            if col_idx is None:
                return None
            
            target_date_only = target_date.date() if isinstance(target_date, datetime) else target_date
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[col_idx]:
                    date_val = row[0]
                    if isinstance(date_val, str):
                        try:
                            date_val = datetime.strptime(date_val, '%Y-%m-%d').date()
                        except:
                            continue
                    elif isinstance(date_val, datetime):
                        date_val = date_val.date()
                    
                    if date_val == target_date_only:
                        return float(row[col_idx])
            
            return None
            
        except Exception as e:
            print(f"获取期货价格失败: {e}")
            return None
    
    def _match_trade_record(self, row, strategy, business_type):
        """匹配交易记录
        
        交易记录Excel列说明：
        - row[5]: 期货合约
        - row[6]: 期货方向
        - row[22]: 现货品种（期货-现货/期货-远期）
        - row[4]: 品种1（远期-现货交易记录）
        - row[10]: 品种2（远期-现货交易记录，K列）
        """
        futures_contract = strategy.get('futures_contract', '')
        spot_variety = strategy.get('spot_variety', '')
        futures_direction = strategy.get('futures_direction', '')
        
        if business_type in ['期货-现货', '期货-远期']:
            # 期货-现货/期货-远期：匹配期货合约、现货/远期品种、期货方向
            return (
                str(row[5]).upper() == str(futures_contract).upper() and
                str(row[22]) == str(spot_variety) and
                str(row[6]) == str(futures_direction)
            )
        elif business_type == '期货-期货':
            # 期货-期货：匹配期货合约1或期货合约2，同时匹配方向
            contract2 = strategy.get('futures_contract2', '')
            return (
                (str(row[5]).upper() == str(futures_contract).upper() or
                 str(row[5]).upper() == str(contract2).upper()) and
                str(row[6]) == str(futures_direction)
            )
        elif business_type == '远期-现货':
            # 远期-现货（两个现货品种价差）：必须同时匹配品种1和品种2
            # 品种1在Excel D列（row[4]），品种2在Excel K列（row[10]）
            variety2 = strategy.get('forward_variety', '')
            row_variety1 = str(row[4]) if row[4] else ''  # 品种1
            row_variety2 = str(row[10]) if row[10] else ''  # 品种2（K列）
            
            # 同时匹配品种1和品种2（顺序可以互换）
            variety_match = (
                (row_variety1 == spot_variety and row_variety2 == variety2) or
                (row_variety1 == variety2 and row_variety2 == spot_variety)
            )
            
            return variety_match and str(row[6]) == str(futures_direction)
        elif business_type == '投机':
            # 投机：只匹配期货合约
            return str(row[5]).upper() == str(futures_contract).upper()
        
        return False
    
    def get_strategy_weekly_data(self, strategy, all_trade_records=None):
        """获取策略按周复盘数据 - 优化版：支持缓存交易记录"""
        weekly_data = []
        
        # 如果没有提供缓存的交易记录，则自己读取
        if all_trade_records is None:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            business_type = strategy.get('business_type', '')
            if business_type == '远期-现货':
                cursor.execute("SELECT file_path FROM data_files WHERE file_type = 'forward_spot_record' ORDER BY upload_date DESC LIMIT 1")
            else:
                cursor.execute("SELECT file_path FROM data_files WHERE file_type = 'trade_record' ORDER BY upload_date DESC LIMIT 1")
            result = cursor.fetchone()
            conn.close()
            
            if result and os.path.exists(result[0]):
                try:
                    wb = openpyxl.load_workbook(result[0], data_only=True)
                    ws = wb.active
                    if business_type == '远期-现货':
                        all_trade_records = list(ws.iter_rows(min_row=4, values_only=True))
                    else:
                        all_trade_records = list(ws.iter_rows(min_row=3, values_only=True))
                except Exception as e:
                    print(f"读取交易记录失败: {e}")
                    return weekly_data
            else:
                return weekly_data
        
        try:
            strategy_type = strategy.get('strategy_type', '')
            business_type = strategy.get('business_type', '')
            futures_direction = strategy.get('futures_direction', '')
            direction_factor = -1 if futures_direction == '空' else 1
            
            weeks = {}
            
            # 获取策略的合约/品种信息用于计算基差
            futures_contract = strategy.get('futures_contract', '')
            spot_variety = strategy.get('spot_variety', '')
            
            for row in all_trade_records:
                # 根据业务类型确定列索引
                if business_type == '远期-现货':
                    # 远期-现货：row[0]是策略类型，row[2]是日期
                    row_strategy_type = str(row[0]) if row[0] else ''
                    row_date = row[2]
                    
                    if row_strategy_type != strategy_type:
                        continue
                    
                    match = self._match_trade_record(row, strategy, business_type)
                    if not match or not row_date:
                        continue
                    
                    # 远期-现货的列映射
                    # row[7]: 价格1, row[8]: 吨数1, row[13]: 价格2, row[14]: 吨数2
                    # row[22]: 平仓价格1, row[23]: 平仓吨数1, row[28]: 平仓价格2
                    open_tons = float(row[8]) if row[8] else 0
                    open_price1 = float(row[7]) if row[7] else 0
                    open_price2 = float(row[13]) if row[13] else 0
                    close_tons = float(row[23]) if len(row) > 23 and row[23] else 0
                    close_price1 = float(row[22]) if len(row) > 22 and row[22] else 0
                    close_price2 = float(row[28]) if len(row) > 28 and row[28] else 0
                    
                    # 转换为手数（假设1手=10吨）
                    open_lots = open_tons / 10
                    close_lots = close_tons / 10
                    
                    # 计算开仓基差（价差）
                    open_basis = open_price1 - open_price2
                    close_basis = None
                    if close_tons > 0:
                        close_basis = close_price1 - close_price2
                    
                else:
                    # 期货类策略：row[1]是策略类型，row[4]是日期
                    row_strategy_type = str(row[1]) if row[1] else ''
                    row_date = row[4]
                    
                    if row_strategy_type != strategy_type:
                        continue
                    
                    match = self._match_trade_record(row, strategy, business_type)
                    if not match or not row_date:
                        continue
                    
                    # 期货类的列映射
                    open_lots = float(row[7]) if row[7] else 0
                    close_lots = float(row[14]) if row[14] else 0
                    open_price = float(row[8]) if row[8] else 0
                    close_price = float(row[15]) if row[15] else 0
                    spot_price = float(row[24]) if row[24] else 0
                    
                    # 计算基差
                    open_basis = spot_price - open_price if spot_price and open_price else None
                    close_basis = None
                    if close_lots > 0:
                        close_spot = float(row[29]) if row[29] else spot_price
                        close_basis = close_spot - close_price if close_spot and close_price else None
                
                # 计算周次
                if isinstance(row_date, datetime):
                    week_num = row_date.isocalendar()[1]
                    week_key = f"{row_date.year}年第{week_num}周"
                else:
                    week_key = "未知周次"
                
                if week_key not in weeks:
                    weeks[week_key] = {
                        'week': week_key,
                        'open_lots': 0,
                        'close_lots': 0,
                        'close_pnl': 0,
                        'position_lots': 0,
                        'open_basis_sum': 0,
                        'open_basis_count': 0,
                        'close_basis_sum': 0,
                        'close_basis_count': 0,
                    }
                
                weeks[week_key]['open_lots'] += open_lots
                weeks[week_key]['close_lots'] += close_lots
                
                # 累计基差数据用于计算平均
                if open_basis is not None:
                    weeks[week_key]['open_basis_sum'] += open_basis
                    weeks[week_key]['open_basis_count'] += 1
                if close_basis is not None:
                    weeks[week_key]['close_basis_sum'] += close_basis
                    weeks[week_key]['close_basis_count'] += 1
                
                # 计算平仓盈亏
                if close_lots > 0:
                    if business_type == '远期-现货':
                        # 价差交易盈亏计算
                        open_spread = open_price1 - open_price2
                        close_spread = close_price1 - close_price2
                        spread_change = close_spread - open_spread
                        if futures_direction == '空':
                            pnl = -spread_change * close_tons
                        else:
                            pnl = spread_change * close_tons
                    else:
                        # 期货类盈亏计算
                        open_price_futures = float(row[8]) if row[8] else 0
                        close_price_futures = float(row[15]) if row[15] else 0
                        pnl = (close_price_futures - open_price_futures) * close_lots * 10 * direction_factor
                    
                    weeks[week_key]['close_pnl'] += pnl
            
            # 计算累计持仓和期末基差
            cumulative_position = 0
            prev_cumulative = 0
            sorted_weeks = sorted(weeks.keys())
            
            for i, week_key in enumerate(sorted_weeks):
                weeks[week_key]['position_lots'] = weeks[week_key]['open_lots'] - weeks[week_key]['close_lots']
                cumulative_position += weeks[week_key]['position_lots']
                weeks[week_key]['cumulative_position'] = cumulative_position
                
                # 计算平均基差
                if weeks[week_key]['open_basis_count'] > 0:
                    weeks[week_key]['avg_open_basis'] = weeks[week_key]['open_basis_sum'] / weeks[week_key]['open_basis_count']
                else:
                    weeks[week_key]['avg_open_basis'] = None
                
                if weeks[week_key]['close_basis_count'] > 0:
                    weeks[week_key]['close_basis'] = weeks[week_key]['close_basis_sum'] / weeks[week_key]['close_basis_count']
                else:
                    weeks[week_key]['close_basis'] = None
                
                # 计算期末基差（取该周最后一天的基差或最新价格计算的基差）
                weeks[week_key]['end_basis'] = self._calculate_week_end_basis(strategy, week_key)
                
                # 计算持仓盈亏（基于期末基差）
                if cumulative_position > 0 and weeks[week_key]['end_basis'] is not None:
                    weeks[week_key]['position_pnl'] = self._calculate_week_position_pnl(
                        strategy, cumulative_position, weeks[week_key]['end_basis'], prev_cumulative
                    )
                else:
                    weeks[week_key]['position_pnl'] = 0
                
                prev_cumulative = cumulative_position
                
                # 转换为吨位 (1手 = 10吨)
                weeks[week_key]['open_tons'] = weeks[week_key]['open_lots'] * 10
                weeks[week_key]['close_tons'] = weeks[week_key]['close_lots'] * 10
                weeks[week_key]['cumulative_tons'] = cumulative_position * 10
                
                # 清理临时数据
                del weeks[week_key]['open_basis_sum']
                del weeks[week_key]['open_basis_count']
                del weeks[week_key]['close_basis_sum']
                del weeks[week_key]['close_basis_count']
            
            weekly_data = list(weeks.values())
            
        except Exception as e:
            print(f"处理交易记录失败: {e}")
            import traceback
            traceback.print_exc()
        
        return weekly_data
    
    def _calculate_week_end_basis(self, strategy, week_key):
        """计算该周期末基差"""
        try:
            business_type = strategy.get('business_type', '')
            
            if business_type == '远期-现货':
                variety1 = strategy.get('spot_variety', '')
                variety2 = strategy.get('forward_variety', '')
                
                variety1_info = self._get_latest_spot_price_with_date(variety1)
                if not variety1_info:
                    return None
                
                current_price1 = variety1_info['price']
                latest_date = variety1_info['date']
                
                current_price2 = self._get_spot_price_by_date(variety2, latest_date)
                if current_price2 is None:
                    return None
                
                return round(current_price1 - current_price2, 2)
            
            else:
                # 期货类策略
                futures_contract = strategy.get('futures_contract', '')
                spot_variety = strategy.get('spot_variety', '')
                
                futures_info = self._get_latest_futures_price_with_date(futures_contract)
                if not futures_info:
                    return None
                
                current_futures_price = futures_info['price']
                latest_date = futures_info['date']
                
                current_spot_price = self._get_spot_price_by_date(spot_variety, latest_date)
                if current_spot_price is None:
                    return None
                
                return round(current_spot_price - current_futures_price, 2)
        
        except Exception as e:
            print(f"计算期末基差失败: {e}")
            return None
    
    def _calculate_week_position_pnl(self, strategy, cumulative_position, end_basis, prev_cumulative):
        """计算该周持仓盈亏"""
        try:
            business_type = strategy.get('business_type', '')
            futures_direction = strategy.get('futures_direction', '')
            
            # 简化计算：假设建仓基差为0，实际应根据持仓明细计算
            # 这里使用期末基差 * 持仓手数 * 10吨/手 * 方向
            if business_type == '远期-现货':
                # 价差交易
                if futures_direction == '空':
                    return round(-end_basis * cumulative_position * 10, 2)
                else:
                    return round(end_basis * cumulative_position * 10, 2)
            else:
                # 期货类
                if futures_direction == '空':
                    return round(end_basis * cumulative_position * 10, 2)
                else:
                    return round(-end_basis * cumulative_position * 10, 2)
        
        except Exception as e:
            print(f"计算持仓盈亏失败: {e}")
            return 0
    
    def get_all_strategies_weekly_data(self, strategies):
        """批量获取多个策略的按周复盘数据 - 优化版：只读取一次Excel"""
        # 一次性读取所有交易记录
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT file_path FROM data_files WHERE file_type = 'trade_record' ORDER BY upload_date DESC LIMIT 1")
        result = cursor.fetchone()
        conn.close()
        
        all_trade_records = []
        if result and os.path.exists(result[0]):
            try:
                wb = openpyxl.load_workbook(result[0], data_only=True)
                ws = wb.active
                all_trade_records = list(ws.iter_rows(min_row=3, values_only=True))
            except Exception as e:
                print(f"读取交易记录失败: {e}")
        
        # 为每个策略计算复盘数据
        result = {}
        for strategy in strategies:
            weekly_data = self.get_strategy_weekly_data(strategy, all_trade_records)
            result[strategy['id']] = weekly_data
        
        return result
