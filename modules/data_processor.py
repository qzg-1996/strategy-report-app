"""
数据处理器 - 处理Excel文件数据
"""
import openpyxl
import sqlite3
import os
from datetime import datetime

class DataProcessor:
    def __init__(self):
        self.db_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'strategy.db')
    
    def process_file(self, file_type, filepath):
        """处理上传的文件"""
        if file_type == 'spot_price':
            return self.process_spot_price(filepath)
        elif file_type == 'futures_price':
            return self.process_futures_price(filepath)
        elif file_type == 'trade_record':
            return self.process_trade_record(filepath)
        elif file_type == 'forward_spot_record':
            return self.process_forward_spot_record(filepath)
        else:
            raise ValueError(f'未知的文件类型: {file_type}')
    
    def process_spot_price(self, filepath):
        """处理现货价格表"""
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        data = []
        headers = []
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:
                headers = row
            else:
                if row[0]:  # 有日期
                    row_data = {}
                    for i, header in enumerate(headers):
                        if header:
                            row_data[header] = row[i] if i < len(row) else None
                    data.append(row_data)
        
        return {'headers': headers, 'row_count': len(data), 'sample': data[:5]}
    
    def process_futures_price(self, filepath):
        """处理期货价格表"""
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        data = []
        headers = []
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:
                headers = row
            else:
                if row[0]:  # 有日期
                    row_data = {}
                    for i, header in enumerate(headers):
                        if header:
                            row_data[header] = row[i] if i < len(row) else None
                    data.append(row_data)
        
        return {'headers': headers, 'row_count': len(data), 'sample': data[:5]}
    
    def process_trade_record(self, filepath):
        """处理交易记录1"""
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        records = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row[1]:  # 策略类型为空，跳过
                continue
            
            record = {
                'row_num': row_idx,
                'strategy_type': row[1],  # 策略类型
                'business_type': row[2],  # 业务类型
                'customer': row[3],  # 客户
                'open_date': row[4],  # 建仓日期
                'futures_contract': row[5],  # 期货合约
                'futures_direction': row[6],  # 期货方向
                'futures_lots': row[7],  # 期货手数
                'futures_open_price': row[8],  # 期货建仓价
                'futures_tons': row[9],  # 期货吨数
                'futures_open_amount': row[10],  # 期货建仓金额
                'close_date': row[11],  # 平仓日期
                'close_contract': row[12],  # 平仓合约
                'close_direction': row[13],  # 平仓方向
                'close_lots': row[14],  # 平仓手数
                'futures_close_price': row[15],  # 期货平仓价
                'close_tons': row[16],  # 平仓吨数
                'profit_points': row[17],  # 盈亏点数
                'futures_profit': row[18],  # 期货盈亏金额
                'total_tons': row[19],  # 总吨数
                'total_profit': row[20],  # 总盈亏
                'spot_open_date': row[21],  # 现货建仓日期
                'spot_variety': row[22],  # 现货品种
                'spot_tons': row[23],  # 现货吨数
                'spot_open_price': row[24],  # 现货建仓价
                'spot_open_amount': row[25],  # 现货建仓金额
                'contract_no': row[26],  # 合同号
                'spot_close_date': row[27],  # 现货平仓日期
                'spot_close_tons': row[28],  # 现货平仓吨数
                'spot_close_price': row[29],  # 现货平仓价
                'spot_close_amount': row[30],  # 现货平仓金额
                'spot_profit': row[32],  # 现货盈亏
                'strategy_total_profit': row[33],  # 策略总盈亏
            }
            records.append(record)
        
        return {
            'row_count': len(records),
            'strategy_types': list(set(r['strategy_type'] for r in records if r['strategy_type'])),
            'business_types': list(set(r['business_type'] for r in records if r['business_type'])),
            'sample': records[:3]
        }
    
    def process_forward_spot_record(self, filepath):
        """处理远期-现货交易记录"""
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        
        records = []
        
        # 表头在第3行，数据从第4行开始
        headers = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), 1):
            if row_idx == 3:
                headers = row
            elif row_idx >= 4 and row[0]:  # 有策略类型
                record = {}
                for i, header in enumerate(headers):
                    if header and i < len(row):
                        record[header] = row[i]
                records.append(record)
        
        return {
            'headers': headers,
            'row_count': len(records),
            'sample': records[:3]
        }
    
    def calculate_strategy_profit(self, strategy_type, business_type, **kwargs):
        """
        计算策略盈亏
        
        参数根据业务类型不同：
        - 期货-现货/期货-远期: futures_contract, futures_direction, spot_variety
        - 期货-期货: spread_type (卷螺差/远近差/螺矿比)
        - 远期-现货: variety1, variety2
        """
        if strategy_type in ['套期保值', '基差交易'] and business_type in ['期货-现货', '期货-远期']:
            return self.calculate_futures_spot_profit(**kwargs)
        elif strategy_type == '基差交易' and business_type == '期货-期货':
            return self.calculate_futures_futures_profit(**kwargs)
        elif strategy_type == '基差交易' and business_type == '远期-现货':
            return self.calculate_forward_spot_profit(**kwargs)
        elif strategy_type == '趋势交易':
            return self.calculate_trend_profit(**kwargs)
        else:
            raise ValueError(f'未支持的策略类型: {strategy_type} - {business_type}')
    
    def calculate_futures_spot_profit(self, futures_contract, futures_direction, spot_variety, trade_records):
        """计算期货-现货盈亏"""
        # 筛选匹配的交易记录
        matching_records = [
            r for r in trade_records
            if r['futures_contract'] == futures_contract
            and r['futures_direction'] == futures_direction
            and r['spot_variety'] == spot_variety
        ]
        
        total_futures_profit = sum(r['futures_profit'] for r in matching_records if r['futures_profit'])
        total_spot_profit = sum(r['spot_profit'] for r in matching_records if r['spot_profit'])
        total_profit = total_futures_profit + total_spot_profit
        
        return {
            'futures_profit': total_futures_profit,
            'spot_profit': total_spot_profit,
            'total_profit': total_profit,
            'record_count': len(matching_records)
        }
    
    def calculate_futures_futures_profit(self, spread_type, trade_records):
        """计算期货-期货（价差）盈亏 - 汇总所有同名合约"""
        # 筛选匹配的交易记录（业务类型=卷螺差/远近差/螺矿比）
        matching_records = [
            r for r in trade_records
            if r['business_type'] == spread_type
        ]
        
        total_profit = sum(r['futures_profit'] for r in matching_records if r['futures_profit'])
        
        return {
            'total_profit': total_profit,
            'record_count': len(matching_records)
        }
    
    def calculate_forward_spot_profit(self, variety1, variety2, trade_records):
        """计算远期-现货盈亏"""
        # 筛选匹配的交易记录
        matching_records = [
            r for r in trade_records
            if r.get('品种1') == variety1 and r.get('品种2') == variety2
        ]
        
        total_profit = sum(r.get('平仓盈亏', 0) for r in matching_records if r.get('平仓盈亏'))
        
        return {
            'total_profit': total_profit,
            'record_count': len(matching_records)
        }
    
    def calculate_trend_profit(self, trade_records):
        """计算趋势交易盈亏 - 全部汇总"""
        matching_records = [
            r for r in trade_records
            if r['strategy_type'] == '趋势交易'
        ]
        
        total_profit = sum(r['futures_profit'] for r in matching_records if r['futures_profit'])
        
        return {
            'total_profit': total_profit,
            'record_count': len(matching_records)
        }
