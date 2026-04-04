# -*- coding: utf-8 -*-
"""
基差走势图生成器 - 生成matplotlib图表图片
"""
import matplotlib
matplotlib.use('Agg')  # 使用非交互式后端
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime, timedelta
import sqlite3
import os
import openpyxl

class BasisChartGenerator:
    def __init__(self):
        self.db_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'strategy.db')
        # 设置中文字体
        from matplotlib import font_manager
        
        # 查找可用的中文字体文件
        chinese_font_path = None
        font_paths = [
            'C:/Windows/Fonts/simhei.ttf',
            'C:/Windows/Fonts/simsun.ttc',
            'C:/Windows/Fonts/msyh.ttc',
            'C:/Windows/Fonts/msyhbd.ttc',
        ]
        
        for fp in font_paths:
            if os.path.exists(fp):
                chinese_font_path = fp
                break
        
        if chinese_font_path:
            # 创建字体属性
            self.chinese_font = font_manager.FontProperties(fname=chinese_font_path)
        else:
            self.chinese_font = None
        
        plt.rcParams['axes.unicode_minus'] = False
    
    def generate_basis_chart(self, strategy, output_path, days=90):
        """
        生成基差走势图
        
        Args:
            strategy: 策略字典
            output_path: 输出图片路径
            days: 显示多少天的数据
        """
        try:
            # 获取基差数据
            basis_data = self._get_basis_data(strategy, days)
            
            if not basis_data:
                return False
            
            # 创建图表 - 去掉标题，减小高度
            fig, ax = plt.subplots(figsize=(14, 4.5), dpi=120)
            
            dates = [d['date'] for d in basis_data]
            basis_values = [d['basis'] for d in basis_data]
            
            # 使用索引作为x轴，实现连续显示（去掉空白日期）
            x_indices = list(range(len(basis_values)))
            
            # 绘制基差柱状图
            colors = ['#1976d2' if v >= 0 else '#d32f2f' for v in basis_values]
            ax.bar(x_indices, basis_values, color=colors, width=0.7, alpha=0.8)
            
            # 计算并标记当前基差（最新值）
            current_basis = basis_values[-1] if basis_values else 0
            current_index = x_indices[-1] if x_indices else 0
            
            # 用红色虚线标记当前基差
            ax.axhline(y=current_basis, color='r', linestyle='--', linewidth=1.5)
            
            # 标记当前点
            ax.scatter([current_index], [current_basis], color='red', s=50, zorder=5)
            
            # 只设置坐标轴标签，不设置标题
            if self.chinese_font:
                ax.set_xlabel('日期', fontsize=11, fontproperties=self.chinese_font, labelpad=10)
                ax.set_ylabel('基差（元/吨）', fontsize=11, fontproperties=self.chinese_font, labelpad=10)
            else:
                ax.set_xlabel('Date', fontsize=11, labelpad=10)
                ax.set_ylabel('Basis (Yuan/Ton)', fontsize=11, labelpad=10)
            
            # 设置x轴刻度 - 选择性地显示日期标签
            date_count = len(dates)
            if date_count > 60:
                step = max(1, date_count // 12)  # 显示约12个标签
            elif date_count > 30:
                step = max(1, date_count // 10)  # 显示约10个标签
            else:
                step = max(1, date_count // 8)   # 显示约8个标签
            
            # 设置x轴刻度位置
            tick_positions = list(range(0, date_count, step))
            if tick_positions[-1] != date_count - 1:
                tick_positions.append(date_count - 1)
            
            # 格式化日期标签
            tick_labels = []
            for i in tick_positions:
                date_str = dates[i]
                if isinstance(date_str, datetime):
                    tick_labels.append(date_str.strftime('%m/%d'))
                else:
                    tick_labels.append(str(date_str)[5:10].replace('-', '/'))
            
            ax.set_xticks(tick_positions)
            if self.chinese_font:
                ax.set_xticklabels(tick_labels, fontsize=9, fontproperties=self.chinese_font)
            else:
                ax.set_xticklabels(tick_labels, fontsize=9)
            
            # 设置Y轴刻度字体
            if self.chinese_font:
                for label in ax.get_yticklabels():
                    label.set_fontproperties(self.chinese_font)
                    label.set_fontsize(9)
            else:
                ax.tick_params(axis='y', labelsize=9)
            
            # 添加网格
            ax.grid(True, alpha=0.3, linestyle=':', axis='y')
            
            # 不添加图例，当前基差值在图外显示
            
            # 调整布局
            plt.tight_layout(pad=1.0)
            
            # 保存当前基差值到文本文件，供PDF使用
            basis_text_path = output_path.replace('.png', '_basis.txt')
            try:
                with open(basis_text_path, 'w', encoding='utf-8') as f:
                    f.write(f'{current_basis:.2f}')
            except:
                pass
            
            # 保存图片
            plt.savefig(output_path, dpi=150, bbox_inches='tight', 
                       facecolor='white', edgecolor='none')
            plt.close()
            
            return True
            
        except Exception as e:
            print(f"生成基差走势图失败: {e}")
            return False
    
    def _get_basis_data(self, strategy, days=90):
        """获取基差数据"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            futures_contract = strategy.get('futures_contract', '')
            spot_variety = strategy.get('spot_variety', '')
            
            # 读取期货价格
            cursor.execute("SELECT file_path FROM data_files WHERE file_type = 'futures_price' ORDER BY upload_date DESC LIMIT 1")
            futures_file = cursor.fetchone()
            
            # 读取现货价格
            cursor.execute("SELECT file_path FROM data_files WHERE file_type = 'spot_price' ORDER BY upload_date DESC LIMIT 1")
            spot_file = cursor.fetchone()
            conn.close()
            
            if not futures_file or not spot_file:
                return None
            
            # 读取数据
            futures_prices = self._read_price_data(futures_file[0], futures_contract)
            
            # 现货品种直接对应现货表中的列名
            spot_col = spot_variety
            spot_prices = self._read_price_data(spot_file[0], spot_col)
            
            # 计算基差：基差 = 期货价 - 现货价（卖-买）
            basis_data = []
            for date in sorted(set(futures_prices.keys()) & set(spot_prices.keys())):
                if futures_prices[date] and spot_prices[date]:
                    basis = futures_prices[date] - spot_prices[date]
                    basis_data.append({
                        'date': date,
                        'basis': round(basis, 2),
                        'spot_price': spot_prices[date],
                        'futures_price': futures_prices[date]
                    })
            
            # 只取最近days天
            if len(basis_data) > days:
                basis_data = basis_data[-days:]
            
            return basis_data
            
        except Exception as e:
            print(f"获取基差数据失败: {e}")
            return None
    
    def _read_price_data(self, file_path, column_name):
        """读取价格数据"""
        prices = {}
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            col_idx = None
            for i, h in enumerate(headers):
                if h == column_name:
                    col_idx = i
                    break
            
            # 如果没找到，尝试主链
            if col_idx is None and '主链' not in column_name:
                main_names = ['螺纹主链', '热卷主链', 'RB主链', 'HC主链']
                for main in main_names:
                    for i, h in enumerate(headers):
                        if h == main:
                            col_idx = i
                            break
                    if col_idx is not None:
                        break
            
            if col_idx is None:
                return prices
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[col_idx]:
                    date_val = row[0]
                    if isinstance(date_val, str):
                        try:
                            date_val = datetime.strptime(date_val, '%Y-%m-%d')
                        except:
                            continue
                    
                    try:
                        price = float(row[col_idx])
                        prices[date_val] = price
                    except:
                        pass
            
            return prices
            
        except Exception as e:
            print(f"读取价格数据失败: {e}")
            return prices
