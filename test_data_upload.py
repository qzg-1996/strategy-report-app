# -*- coding: utf-8 -*-
"""
测试数据上传功能
"""
import openpyxl
import sqlite3
import os

def test_spot_price():
    """测试现货价格数据"""
    file_path = r'C:\Users\Administrator\Desktop\现货价格 - 副本(1)(1).xlsx'
    print(f"\n=== 测试现货价格表: {file_path} ===")
    
    if not os.path.exists(file_path):
        print("文件不存在!")
        return None
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # 读取表头
        headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        print(f"列名: {headers}")
        print(f"总行数: {ws.max_row}")
        
        # 读取前3行数据
        print("\n前3行数据:")
        for row in ws.iter_rows(min_row=2, max_row=4, values_only=True):
            print(row[:5])
        
        wb.close()
        return headers
    except Exception as e:
        print(f"读取失败: {e}")
        return None

def test_futures_price():
    """测试期货价格数据"""
    file_path = r'C:\Users\Administrator\Desktop\期货价格 - 副本(1)(1).xlsx'
    print(f"\n=== 测试期货价格表: {file_path} ===")
    
    if not os.path.exists(file_path):
        print("文件不存在!")
        return None
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # 读取表头
        headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        print(f"列名: {headers}")
        print(f"总行数: {ws.max_row}")
        
        # 读取前3行数据
        print("\n前3行数据:")
        for row in ws.iter_rows(min_row=2, max_row=4, values_only=True):
            print(row)
        
        wb.close()
        return headers
    except Exception as e:
        print(f"读取失败: {e}")
        return None

def test_trade_record():
    """测试交易记录数据"""
    file_path = r'C:\Users\Administrator\Desktop\交易记录1.xlsx'
    print(f"\n=== 测试交易记录: {file_path} ===")
    
    if not os.path.exists(file_path):
        print("文件不存在!")
        return None
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        print(f"总行数: {ws.max_row}")
        
        # 读取表头（第2行）
        headers = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        print(f"列名(第2行): {headers[:10]}")
        
        # 读取前3条数据（从第3行开始）
        print("\n前3条交易记录:")
        for i, row in enumerate(ws.iter_rows(min_row=3, max_row=5, values_only=True), 3):
            if row[1]:  # 有策略类型
                print(f"Row {i}: 策略={row[1]}, 业务={row[2]}, 合约={row[5]}, 方向={row[6]}, 手数={row[7]}")
        
        wb.close()
        return True
    except Exception as e:
        print(f"读取失败: {e}")
        import traceback
        traceback.print_exc()
        return None

def test_forward_spot():
    """测试远期-现货交易记录"""
    file_path = r'C:\Users\Administrator\Desktop\远期-现货交易记录(1).xlsx'
    print(f"\n=== 测试远期-现货交易记录: {file_path} ===")
    
    if not os.path.exists(file_path):
        print("文件不存在!")
        return None
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        print(f"总行数: {ws.max_row}")
        
        # 读取前3行
        print("\n前3行数据:")
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            print(row[:8])
        
        wb.close()
        return True
    except Exception as e:
        print(f"读取失败: {e}")
        return None

def test_database():
    """测试数据库连接"""
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'strategy.db')
    print(f"\n=== 测试数据库: {db_path} ===")
    
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # 检查表是否存在
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = cursor.fetchall()
        print(f"数据库表: {[t[0] for t in tables]}")
        
        # 检查策略数量
        cursor.execute("SELECT COUNT(*) FROM strategies")
        count = cursor.fetchone()[0]
        print(f"策略数量: {count}")
        
        # 检查数据文件记录
        cursor.execute("SELECT * FROM data_files")
        files = cursor.fetchall()
        print(f"已上传文件: {len(files)}")
        for f in files:
            print(f"  - {f}")
        
        conn.close()
        return True
    except Exception as e:
        print(f"数据库错误: {e}")
        return None

if __name__ == '__main__':
    print("开始测试数据文件...")
    
    spot_headers = test_spot_price()
    futures_headers = test_futures_price()
    test_trade_record()
    test_forward_spot()
    test_database()
    
    print("\n=== 测试完成 ===")
