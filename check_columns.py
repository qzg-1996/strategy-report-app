# -*- coding: utf-8 -*-
import openpyxl
import sys

if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\Administrator\Desktop\交易记录1.xlsx', data_only=True)
ws = wb.active

print('=== 第1行表头 ===')
row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
for i, v in enumerate(row1[:25]):
    print(f'列{i}: {v}')

print('\n=== 第2行表头 ===')
row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
for i, v in enumerate(row2[:25]):
    print(f'列{i}: {v}')

print('\n=== 找一列套期保值的数据行（行5）详细查看 ===')
row5 = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))[0]
for i, v in enumerate(row5[:35]):
    print(f'列{i}: {v}')
